from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from build_by_tz import TZ_DOCX, build_corpus, extract_text_from_docx
from recount_semantic_codes import MACRO_CODES, code_macro

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "02_Пересчет_смысловых_кодов"
OUT_XLSX = OUT / "04_Отчет_по_тональности.xlsx"

TONE_META: tuple[tuple[str, str], ...] = (
    ("ТОН_ПОЗИТИВ", "Позитив"),
    ("ТОН_КОНСТРУКТИВ", "Конструктив"),
    ("ТОН_НЕГАТИВ", "Негатив"),
    ("ТОН_СОМНЕНИЕ", "Сомнение"),
)

TONE_POSITIVE = (
    r"мер достаточно",
    r"достаточно мер",
    r"доступ есть",
    r"вс[её] устраивает",
    r"вс[её] нормально",
    r"нет проблем",
    r"не требуется",
    r"всего вполне",
    r"всем вс[её] доступно",
    r"вс[её] доступно",
    r"вс[её] есть",
    r"принимаются достаточные",
)
TONE_DOUBT = (
    r"затрудн",
    r"не знаю",
    r"неизвестно",
    r"надо подумать",
    r"правительству виднее",
    r"пусть решают",
    r"нет предлож",
)
TONE_NEGATIVE = (
    r"невыносим",
    r"выжив",
    r"закрыт",
    r"обидно",
    r"монопол",
    r"задерж",
    r"кассов",
    r"убыт",
    r"штраф",
    r"шантаж",
    r"не работает",
    r"искусственн",
    r"не поможет",
    r"ничего не поможет",
    r"руки опуск",
    r"низк.*тариф|тариф.*низк",
    r"не индекс",
    r"остаточн",
    r"формально",
)
PROPOSAL_MARKERS = (
    r"упрост",
    r"повыс",
    r"увелич",
    r"внедр",
    r"расшир",
    r"обеспеч",
    r"создат",
    r"пересмотр",
    r"индекс",
    r"компенсац",
    r"отмен",
)


def assign_tone(text: str, semantic: dict[str, int]) -> str:
    low = text.lower().strip()
    for pat in TONE_DOUBT:
        if re.search(pat, low):
            return "ТОН_СОМНЕНИЕ"
    for pat in TONE_POSITIVE:
        if re.search(pat, low):
            if not re.search(r"но\s|тариф.*низк|невыносим|задерж", low):
                return "ТОН_ПОЗИТИВ"
    for pat in TONE_NEGATIVE:
        if re.search(pat, low):
            return "ТОН_НЕГАТИВ"
    if any(re.search(p, low) for p in PROPOSAL_MARKERS) or any(semantic.values()):
        return "ТОН_КОНСТРУКТИВ"
    if len(low) < 35:
        return "ТОН_СОМНЕНИЕ"
    return "ТОН_КОНСТРУКТИВ"


def header_row(ws, values: list[str], fill: str = "1F4E79") -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)


def short_text(text: str, limit: int = 190) -> str:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - 3] + "..."


def collect_report_data() -> dict[str, object]:
    raw = extract_text_from_docx(TZ_DOCX)
    statements, rough = build_corpus(raw)

    rows: list[dict[str, object]] = []
    tone_totals = {key: 0 for key, _ in TONE_META}
    semantic_totals = {macro.key: 0 for macro in MACRO_CODES}
    cross: dict[str, dict[str, int]] = {
        macro.key: {tone_key: 0 for tone_key, _ in TONE_META} for macro in MACRO_CODES
    }
    examples_by_tone = {tone_key: [] for tone_key, _ in TONE_META}

    for stmt in statements:
        semantic = code_macro(stmt.text)
        tone = assign_tone(stmt.text, semantic)
        tone_totals[tone] += 1

        if len(examples_by_tone[tone]) < 5:
            examples_by_tone[tone].append(short_text(stmt.text))

        for macro in MACRO_CODES:
            if semantic[macro.key]:
                semantic_totals[macro.key] += 1
                cross[macro.key][tone] += 1

        rows.append(
            {
                "id": stmt.id,
                "parent_id": stmt.parent_id,
                "text": stmt.text,
                "tone": tone,
                **semantic,
            }
        )

    total_n = len(rows)
    macro_by_key = {macro.key: macro for macro in MACRO_CODES}
    tone_labels = {key: label for key, label in TONE_META}
    ranked = sorted(semantic_totals.items(), key=lambda item: (-item[1], item[0]))

    cross_rows: list[dict[str, object]] = []
    pct_rows: list[dict[str, object]] = []
    for macro in MACRO_CODES:
        counts = cross[macro.key]
        total = sum(counts.values())
        cross_rows.append(
            {
                "key": macro.key,
                "name": macro.name,
                "Позитив": counts["ТОН_ПОЗИТИВ"],
                "Конструктив": counts["ТОН_КОНСТРУКТИВ"],
                "Негатив": counts["ТОН_НЕГАТИВ"],
                "Сомнение": counts["ТОН_СОМНЕНИЕ"],
                "Итого": total,
            }
        )
        pct_rows.append(
            {
                "key": macro.key,
                "name": macro.name,
                "Позитив": counts["ТОН_ПОЗИТИВ"] / total if total else 0,
                "Конструктив": counts["ТОН_КОНСТРУКТИВ"] / total if total else 0,
                "Негатив": counts["ТОН_НЕГАТИВ"] / total if total else 0,
                "Сомнение": counts["ТОН_СОМНЕНИЕ"] / total if total else 0,
            }
        )

    top_by_tone: dict[str, list[tuple[str, int]]] = {}
    for tone_key, label in TONE_META:
        tone_ranked = sorted(
            ((macro.name, cross[macro.key][tone_key]) for macro in MACRO_CODES),
            key=lambda item: (-item[1], item[0]),
        )
        top_by_tone[label] = [item for item in tone_ranked if item[1] > 0][:3]

    return {
        "rows": rows,
        "n_rough": len(rough),
        "n_atomic": total_n,
        "tone_totals": tone_totals,
        "semantic_totals": semantic_totals,
        "cross": cross,
        "cross_rows": cross_rows,
        "pct_rows": pct_rows,
        "examples_by_tone": examples_by_tone,
        "tone_labels": tone_labels,
        "macro_by_key": macro_by_key,
        "ranked": ranked,
        "top_by_tone": top_by_tone,
    }


def build_report() -> Path:
    data = collect_report_data()
    rows = data["rows"]
    total_n = data["n_atomic"]
    tone_totals = data["tone_totals"]
    ranked = data["ranked"]
    tone_labels = data["tone_labels"]

    OUT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "Данные"
    header_row(
        ws_data,
        ["ID", "ID_блока", "Текст", "Тональность"] + [macro.key for macro in MACRO_CODES],
    )
    for row in rows:
        ws_data.append(
            [row["id"], row["parent_id"], row["text"], tone_labels[row["tone"]]]
            + [row[macro.key] for macro in MACRO_CODES]
        )
    ws_data.column_dimensions["C"].width = 90
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws_data.freeze_panes = "D2"
    last = total_n + 1
    code_start = 5

    ws_tone = wb.create_sheet("Итоги_тональность")
    ws_tone["A1"] = "ТОНАЛЬНОСТЬ ВЫСКАЗЫВАНИЙ"
    ws_tone["A1"].font = Font(bold=True, size=14)
    ws_tone.merge_cells("A1:D1")
    ws_tone["A2"] = f"Всего атомарных высказываний: {total_n}"
    ws_tone["A2"].font = Font(bold=True)
    ws_tone.append([])
    header_row(ws_tone, ["Тональность", "Частота", "Доля"], fill="D9E2F3")
    for i, (tone_key, label) in enumerate(TONE_META, 4):
        cnt = tone_totals[tone_key]
        ws_tone.cell(row=i, column=1, value=label)
        ws_tone.cell(row=i, column=2, value=cnt)
        ws_tone.cell(row=i, column=3, value=cnt / total_n if total_n else 0)
        ws_tone.cell(row=i, column=3).number_format = "0.0%"

    pie = PieChart()
    pie.title = "Распределение по тональности"
    pie.height = 10
    pie.width = 14
    pie.add_data(Reference(ws_tone, min_col=2, min_row=3, max_row=3 + len(TONE_META)), titles_from_data=True)
    pie.set_categories(Reference(ws_tone, min_col=1, min_row=4, max_row=3 + len(TONE_META)))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    ws_tone.add_chart(pie, "E3")

    ws_sem = wb.create_sheet("Итоги_смыслы")
    ws_sem["A1"] = "ЧАСТОТА СМЫСЛОВЫХ КАТЕГОРИЙ"
    ws_sem["A1"].font = Font(bold=True, size=14)
    ws_sem.merge_cells("A1:E1")
    ws_sem["A2"] = f"Исходных блоков: {data['n_rough']}; атомарных высказываний: {total_n}"
    ws_sem["A2"].font = Font(bold=True)
    ws_sem.append([])
    header_row(ws_sem, ["Ранг", "Код", "Смысловая категория", "Частота", "Доля"], fill="D9E2F3")

    sem_start = 5
    for rank, (key, _cnt) in enumerate(ranked, 1):
        row_idx = sem_start + rank - 1
        macro = data["macro_by_key"][key]
        col_idx = next(i for i, m in enumerate(MACRO_CODES) if m.key == key)
        col_letter = get_column_letter(code_start + col_idx)
        ws_sem.cell(row=row_idx, column=1, value=rank)
        ws_sem.cell(row=row_idx, column=2, value=key)
        ws_sem.cell(row=row_idx, column=3, value=macro.name)
        ws_sem.cell(row=row_idx, column=4, value=f"=SUM(Данные!{col_letter}2:{col_letter}{last})")
        ws_sem.cell(row=row_idx, column=5, value=f"=D{row_idx}/{total_n}")
        ws_sem.cell(row=row_idx, column=5).number_format = "0.0%"

    ws_sem.column_dimensions["C"].width = 42
    bar = BarChart()
    bar.type = "col"
    bar.title = "Частота смысловых категорий"
    bar.height = 12
    bar.width = 18
    bar.add_data(Reference(ws_sem, min_col=4, min_row=4, max_row=sem_start + len(MACRO_CODES) - 1), titles_from_data=True)
    bar.set_categories(Reference(ws_sem, min_col=3, min_row=5, max_row=sem_start + len(MACRO_CODES) - 1))
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    ws_sem.add_chart(bar, "G4")

    ws_cross = wb.create_sheet("Смыслы_x_Тон")
    ws_cross["A1"] = "СМЫСЛОВЫЕ КАТЕГОРИИ ПО ТОНАЛЬНОСТИ"
    ws_cross["A1"].font = Font(bold=True, size=14)
    ws_cross.merge_cells("A1:G1")
    ws_cross["A2"] = (
        "Ячейка = число высказываний, где одновременно присутствует смысловой код и данная тональность. "
        "Одно высказывание может учитываться в нескольких строках."
    )
    ws_cross.merge_cells("A2:G2")
    ws_cross.append([])
    header_row(ws_cross, ["Смысловая категория"] + [label for _, label in TONE_META] + ["Итого"], fill="D9E2F3")

    cross_start = 5
    for i, item in enumerate(data["cross_rows"]):
        row_idx = cross_start + i
        ws_cross.cell(row=row_idx, column=1, value=item["name"])
        ws_cross.cell(row=row_idx, column=2, value=item["Позитив"])
        ws_cross.cell(row=row_idx, column=3, value=item["Конструктив"])
        ws_cross.cell(row=row_idx, column=4, value=item["Негатив"])
        ws_cross.cell(row=row_idx, column=5, value=item["Сомнение"])
        ws_cross.cell(row=row_idx, column=6, value=item["Итого"])

    total_row = cross_start + len(MACRO_CODES)
    ws_cross.cell(row=total_row, column=1, value="Итого").font = Font(bold=True)
    for j in range(2, 7):
        col_letter = get_column_letter(j)
        ws_cross.cell(row=total_row, column=j, value=f"=SUM({col_letter}{cross_start}:{col_letter}{total_row - 1})")
        ws_cross.cell(row=total_row, column=j).font = Font(bold=True)

    ws_cross.column_dimensions["A"].width = 42
    stacked = BarChart()
    stacked.type = "col"
    stacked.grouping = "stacked"
    stacked.title = "Смысловые категории по тональности"
    stacked.height = 14
    stacked.width = 22
    stacked.add_data(Reference(ws_cross, min_col=2, max_col=5, min_row=4, max_row=total_row - 1), titles_from_data=True)
    stacked.set_categories(Reference(ws_cross, min_col=1, min_row=cross_start, max_row=total_row - 1))
    stacked.dataLabels = DataLabelList()
    stacked.dataLabels.showVal = False
    ws_cross.add_chart(stacked, "H4")

    ws_pct = wb.create_sheet("Смыслы_x_Тон_доли")
    ws_pct["A1"] = "ДОЛИ ТОНАЛЬНОСТИ ВНУТРИ КАЖДОЙ СМЫСЛОВОЙ КАТЕГОРИИ"
    ws_pct["A1"].font = Font(bold=True, size=14)
    ws_pct.merge_cells("A1:F1")
    ws_pct.append([])
    header_row(ws_pct, ["Смысловая категория"] + [label for _, label in TONE_META], fill="D9E2F3")

    pct_start = 4
    for i, item in enumerate(data["pct_rows"]):
        row_idx = pct_start + i
        ws_pct.cell(row=row_idx, column=1, value=item["name"])
        ws_pct.cell(row=row_idx, column=2, value=item["Позитив"])
        ws_pct.cell(row=row_idx, column=3, value=item["Конструктив"])
        ws_pct.cell(row=row_idx, column=4, value=item["Негатив"])
        ws_pct.cell(row=row_idx, column=5, value=item["Сомнение"])
        for j in range(2, 6):
            ws_pct.cell(row=row_idx, column=j).number_format = "0.0%"

    ws_pct.column_dimensions["A"].width = 42
    pct_bar = BarChart()
    pct_bar.type = "bar"
    pct_bar.grouping = "percentStacked"
    pct_bar.title = "Структура тональности внутри смысловых категорий (%)"
    pct_bar.height = 14
    pct_bar.width = 22
    pct_bar.add_data(Reference(ws_pct, min_col=2, max_col=5, min_row=3, max_row=pct_start + len(MACRO_CODES) - 1), titles_from_data=True)
    pct_bar.set_categories(Reference(ws_pct, min_col=1, min_row=pct_start, max_row=pct_start + len(MACRO_CODES) - 1))
    ws_pct.add_chart(pct_bar, "G3")

    wb.save(OUT_XLSX)
    return OUT_XLSX


if __name__ == "__main__":
    path = build_report()
    print(f"Отчёт по смыслам: {path}")
